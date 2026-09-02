import base64

import pytest
from openpyxl import Workbook, load_workbook

from tools.annotate_excel import AnnotationError, annotate_workbook


TINY_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
)


def _make_workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "用例"
    sheet.append(["用例ID", "用例名称", "优先级", "步骤"])
    sheet.append(["TC-001", "登录", "high", "输入账号"])
    sheet.append([None, None, None, "点击登录"])
    sheet.append(["TC-002", "查询", "middle", "打开查询页"])
    sheet.append(["TC-003", "重复名称", "low", "第一处"])
    sheet.append(["TC-004", "中间用例", "low", "分隔"])
    sheet.append(["TC-005", "重复名称", "low", "第二处"])
    workbook.save(path)


def test_backfills_by_id_row_and_name_without_changing_source(tmp_path):
    source = tmp_path / "cases.xlsx"
    output = tmp_path / "out.xlsx"
    _make_workbook(source)

    report = annotate_workbook(
        source,
        {
            "cases": [
                {"case_id": "TC-001", "row": 2, "status": "✅通过", "actual": "两步均完成", "tested_at": "2026-08-25"},
                {"sheet": "用例", "case_name": "查询", "status": "🟡待数据", "result": "缺少行情数据"},
                {"sheet": "用例", "row": 7, "case_id": "TC-005", "status": "❌失败", "actual": "第二处失败"},
                {"sheet": "用例", "case_id": "NOT-FOUND", "status": "⏭跳过"},
            ]
        },
        output,
        generated_at="2026-08-25T10:00:00+08:00",
    )

    assert len(report["matched"]) == 3
    assert len(report["unmatched"]) == 1
    assert source.exists()

    original = load_workbook(source)
    assert original["用例"].max_column == 4

    workbook = load_workbook(output)
    sheet = workbook["用例"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    assert headers[-4:] == [
        "🤖AI状态",
        "🤖AI实测结果",
        "🤖AI证据",
        "🤖AI时间",
    ]
    status_column = headers.index("🤖AI状态") + 1
    actual_column = headers.index("🤖AI实测结果") + 1
    assert sheet.cell(2, status_column).value == "✅通过"
    assert sheet.cell(3, status_column).value is None
    assert sheet.cell(4, status_column).value == "🟡待数据"
    assert sheet.cell(4, actual_column).value == "缺少行情数据"
    assert sheet.cell(7, status_column).value == "❌失败"
    assert "🤖AI自测汇总" in workbook.sheetnames
    assert workbook["🤖AI自测汇总"].cell(4, 2).value == 3
    assert workbook["🤖AI自测汇总"].cell(5, 2).value == 1


def test_embeds_image_evidence_and_is_idempotent(tmp_path):
    source = tmp_path / "cases.xlsx"
    output = tmp_path / "out.xlsx"
    evidence = tmp_path / "evidence.png"
    evidence.write_bytes(TINY_PNG)
    _make_workbook(source)
    result = {
        "cases": [
            {
                "sheet": "用例",
                "row": 4,
                "status": "通过",
                "actual": "ok",
                "evidence": [{"path": "evidence.png", "description": "截图"}],
            }
        ]
    }

    annotate_workbook(source, result, output, result_dir=tmp_path, generated_at="2026-08-25")
    first = load_workbook(output)
    assert len(first["用例"]._images) == 1
    assert first["用例"].cell(4, 7).value == "截图: evidence.png"

    second_output = tmp_path / "out-2.xlsx"
    annotate_workbook(output, result, second_output, result_dir=tmp_path, generated_at="2026-08-25")
    second = load_workbook(second_output)
    assert len(second["用例"]._images) == 1


def test_rejects_ambiguous_case_name_and_accepts_row_disambiguation(tmp_path):
    source = tmp_path / "cases.xlsx"
    _make_workbook(source)

    with pytest.raises(AnnotationError, match="不唯一"):
        annotate_workbook(
            source,
            [{"sheet": "用例", "case_name": "重复名称", "status": "通过"}],
            tmp_path / "ambiguous.xlsx",
        )

    report = annotate_workbook(
        source,
        [{"sheet": "用例", "row": 5, "case_name": "重复名称", "status": "通过"}],
        tmp_path / "resolved.xlsx",
    )
    assert report["matched"][0]["rows"] == [5]


def test_strict_mode_does_not_save_unmatched_results(tmp_path):
    source = tmp_path / "cases.xlsx"
    output = tmp_path / "strict.xlsx"
    _make_workbook(source)

    with pytest.raises(AnnotationError, match="strict 模式"):
        annotate_workbook(source, [{"case_id": "missing", "status": "通过"}], output, strict=True)
    assert not output.exists()


def test_custom_header_and_columns_support_an_arbitrary_layout(tmp_path):
    source = tmp_path / "custom.xlsx"
    output = tmp_path / "custom-out.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "自定义"
    sheet.append(["这是标题说明"])
    sheet.append([])
    sheet.append(["Case Key", "Scenario", "Priority", "Action"])
    sheet.append(["C-001", "导出报告", "P1", "点击导出"])
    sheet.append([None, None, None, "检查文件"])
    workbook.save(source)

    report = annotate_workbook(
        source,
        [{"sheet": "自定义", "row": 4, "case_id": "C-001", "status": "通过", "actual": "导出成功"}],
        output,
        header_row=3,
        case_id_column="A",
        case_name_column="B",
    )

    assert report["matched"][0]["rows"] == [4]
    annotated = load_workbook(output)
    sheet = annotated["自定义"]
    headers = [sheet.cell(3, column).value for column in range(1, sheet.max_column + 1)]
    status_column = headers.index("🤖AI状态") + 1
    assert sheet.cell(4, status_column).value == "通过"
    assert sheet.cell(5, status_column).value is None


def test_case_level_result_does_not_broadcast_to_contiguous_steps(tmp_path):
    source = tmp_path / "cases.xlsx"
    output = tmp_path / "expanded.xlsx"
    _make_workbook(source)

    report = annotate_workbook(
        source,
        [{"sheet": "用例", "row": 2, "case_name": "登录", "status": "样例待执行"}],
        output,
    )

    assert report["matched"][0]["rows"] == [2]
    workbook = load_workbook(output)
    sheet = workbook["用例"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    status_column = headers.index("🤖AI状态") + 1
    assert sheet.cell(2, status_column).value == "样例待执行"
    assert sheet.cell(3, status_column).value is None


def test_step_results_are_written_only_to_their_exact_source_rows(tmp_path):
    source = tmp_path / "cases.xlsx"
    output = tmp_path / "step-results.xlsx"
    _make_workbook(source)

    report = annotate_workbook(
        source,
        {
            "cases": [
                {
                    "sheet": "用例",
                    "row": 2,
                    "case_id": "TC-001",
                    "steps": [
                        {
                            "step_id": "TC-001#S1",
                            "step_index": 1,
                            "row": 2,
                            "status": "✅通过",
                            "actual": "账号输入成功",
                        },
                        {
                            "step_id": "TC-001#S2",
                            "step_index": 2,
                            "row": 3,
                            "status": "❌失败",
                            "actual": "登录按钮无响应",
                        },
                    ],
                }
            ]
        },
        output,
        append_summary=False,
    )

    assert report["matched"][0]["rows"] == [2, 3]
    assert [item["step_id"] for item in report["matched_steps"]] == ["TC-001#S1", "TC-001#S2"]
    workbook = load_workbook(output)
    sheet = workbook["用例"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    status_column = headers.index("🤖AI状态") + 1
    actual_column = headers.index("🤖AI实测结果") + 1
    assert sheet.cell(2, status_column).value == "✅通过"
    assert sheet.cell(2, actual_column).value == "S1 [✅通过] 账号输入成功"
    assert sheet.cell(3, status_column).value == "❌失败"
    assert sheet.cell(3, actual_column).value == "S2 [❌失败] 登录按钮无响应"


def test_multiple_steps_on_one_source_row_are_combined_in_order(tmp_path):
    source = tmp_path / "cases.xlsx"
    output = tmp_path / "same-row-steps.xlsx"
    _make_workbook(source)

    annotate_workbook(
        source,
        {
            "cases": [
                {
                    "sheet": "用例",
                    "row": 2,
                    "case_id": "TC-001",
                    "steps": [
                        {"step_id": "TC-001#S1", "status": "✅通过", "actual": "进入页面"},
                        {"step_id": "TC-001#S2", "status": "✅通过", "actual": "展示数据"},
                    ],
                }
            ]
        },
        output,
        append_summary=False,
    )

    workbook = load_workbook(output)
    sheet = workbook["用例"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    actual_column = headers.index("🤖AI实测结果") + 1
    assert sheet.cell(2, actual_column).value == "S1 [✅通过] 进入页面\nS2 [✅通过] 展示数据"


def test_duplicate_step_id_is_rejected(tmp_path):
    source = tmp_path / "cases.xlsx"
    _make_workbook(source)

    with pytest.raises(AnnotationError, match="步骤结果重复"):
        annotate_workbook(
            source,
            {
                "cases": [
                    {
                        "sheet": "用例",
                        "row": 2,
                        "case_id": "TC-001",
                        "steps": [
                            {"step_id": "duplicate", "status": "通过"},
                            {"step_id": "duplicate", "status": "通过"},
                        ],
                    }
                ]
            },
            tmp_path / "duplicate.xlsx",
            append_summary=False,
        )
