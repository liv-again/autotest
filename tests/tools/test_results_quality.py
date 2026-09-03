import base64

import pytest
from openpyxl import Workbook, load_workbook

from tools.annotate_excel import AnnotationError, annotate_workbook
from tools.build_results import BuildResultsError, build_results
from tools.contracts.validate import validate


TINY_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
)


def _make_workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "用例"
    sheet.append(["用例ID", "用例名称", "优先级", "步骤"])
    sheet.append(["TC-001", "查询", "high", "点击查询"])
    workbook.save(path)


def _strict_case(*, actual, status="✅通过", evidence=None, case_id="TC-001"):
    return {
        "sheet": "用例",
        "row": 2,
        "case_id": case_id,
        "status": status,
        "actual": actual,
        "evidence": evidence or ["ui assertion: 查询页面标题"],
    }


def test_strict_rejects_generic_actual_before_saving(tmp_path):
    source = tmp_path / "cases.xlsx"
    output = tmp_path / "out.xlsx"
    _make_workbook(source)

    with pytest.raises(AnnotationError, match="质量校验"):
        annotate_workbook(
            source,
            [_strict_case(actual="已在国投 App 中执行该行指定的点击操作并保留截图。")],
            output,
            strict=True,
        )
    assert not output.exists()


def test_strict_preserves_not_applicable_status_and_inherits_case_evidence(tmp_path):
    source = tmp_path / "cases.xlsx"
    output = tmp_path / "out.xlsx"
    evidence = tmp_path / "evidence.png"
    evidence.write_bytes(TINY_PNG)
    _make_workbook(source)

    report = annotate_workbook(
        source,
        {
            "cases": [
                {
                    "sheet": "用例",
                    "row": 2,
                    "case_id": "TC-001",
                    "status": "☑不适用",
                    "evidence": [{"path": "evidence.png", "description": "页面截图"}],
                    "steps": [
                        {
                            "step_id": "TC-001#S1",
                            "step_index": 1,
                            "row": 2,
                            "status": "☑不适用",
                            "actual": "页面显示该功能不适用",
                        }
                    ],
                }
            ]
        },
        output,
        result_dir=tmp_path,
        strict=True,
        append_summary=False,
    )

    assert len(report["matched"]) == 1
    workbook = load_workbook(output)
    sheet = workbook["用例"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    status_column = headers.index("🤖AI状态") + 1
    evidence_column = headers.index("🤖AI证据") + 1
    assert sheet.cell(2, status_column).value == "☑不适用"
    assert sheet.cell(2, evidence_column).value == "页面截图: evidence.png"
    assert len(sheet._images) == 1


def test_strict_rejects_reused_actual_across_unrelated_cases(tmp_path):
    source = tmp_path / "cases.xlsx"
    output = tmp_path / "out.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "用例"
    sheet.append(["用例ID", "用例名称", "优先级", "步骤"])
    for case_id, name in (("TC-001", "查询"), ("TC-002", "排序"), ("TC-003", "详情")):
        sheet.append([case_id, name, "high", "操作"])
    workbook.save(source)

    cases = [
        _strict_case(actual="页面状态正常", case_id=case_id)
        for case_id in ("TC-001", "TC-002", "TC-003")
    ]
    for row, case in enumerate(cases, start=2):
        case["row"] = row

    with pytest.raises(AnnotationError, match="复用"):
        annotate_workbook(source, cases, output, strict=True)
    assert not output.exists()


def test_build_results_preserves_setup_trace_and_inherits_evidence():
    document = build_results(
        {
            "cases": [
                {
                    "sheet": "用例",
                    "row": 2,
                    "case_id": "TC-001",
                    "status": "✅通过",
                    "evidence": ["assertions/query-page"],
                    "steps": [
                        {
                            "step_id": "TC-001#S1",
                            "step_index": 1,
                            "row": 2,
                            "status": "✅通过",
                            "actual": "点击查询后进入查询页面并显示结果列表",
                        }
                    ],
                }
            ],
            "setup_trace": [{"kind": "setup", "action": "进入行情首页"}],
        }
    )

    assert document["schema_version"] == "2.0"
    assert document["setup_trace"][0]["kind"] == "setup"
    assert document["cases"][0]["steps"][0]["evidence"] == ["assertions/query-page"]
    assert validate(document, "results") == []


def test_build_results_never_fills_missing_actual():
    with pytest.raises(BuildResultsError, match="actual"):
        build_results(
            {
                "cases": [
                    {
                        "sheet": "用例",
                        "row": 2,
                        "status": "✅通过",
                        "evidence": ["assertions/query-page"],
                    }
                ]
            }
        )
