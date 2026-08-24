"""将结构化自测结果安全地回填到任意 Excel 用例文件。

这个模块不保存任何具体项目、App 或历史批次的数据。LLM 或执行器只需提供
一个 Excel 文件和一个结果 JSON/YAML 文件即可：

.. code-block:: json

   {
     "cases": [
       {
         "sheet": "风险警示需求",
         "row": 2,
         "case_id": "RISK-001",
         "status": "🟡待数据",
         "actual": "当前退市整理列表为空",
         "evidence": ["screenshots/delist-empty.png"],
         "tested_at": "2026-08-25"
       }
     ]
   }

优先使用 ``row`` 或 Excel 中已有的用例 ID 定位。没有 ID 时可以使用
``sheet`` + ``case_name``；如果同名用例分布在多个不连续区块，必须提供
``row``，模块会拒绝猜测。证据图片会嵌入“🤖AI证据”列，非图片证据会以文本
保存。常见表头会自动识别；自定义格式可补充 ``--header-row``、
``--case-id-column``、``--case-name-column``。输出默认写成源文件旁的
``*_AI自测结果.xlsx``，不会覆盖源文件。

CLI 示例：

    python tools/annotate_excel.py --src cases.xlsx --results results.json

也可以直接传入 JSON 字符串，便于 LLM 在一次调用中完成回填：

    python tools/annotate_excel.py --src cases.xlsx --results-json '{"cases": [...]}'
"""

from __future__ import annotations

import argparse
import copy
import json
import os
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import yaml
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string


SUMMARY_SHEET = "🤖AI自测汇总"
OUTPUT_HEADERS = (
    "🤖AI状态",
    "🤖AI实测结果",
    "🤖AI证据",
    "🤖AI时间",
)
LEGACY_HIDDEN_HEADERS = ("🤖AI用例ID", "🤖AI档位")
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tif", ".tiff"}

# 这些只是通用的列名别名，不包含任何具体用例内容。结果中的字段名也支持
# 英文写法，方便 LLM 和不同项目的执行器直接复用。
HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "case_id": (
        "用例id",
        "用例编号",
        "测试用例id",
        "测试用例编号",
        "tcid",
        "tc编号",
        "caseid",
        "case_id",
    ),
    "case_name": (
        "用例名称",
        "用例名",
        "测试用例名称",
        "测试用例名",
        "case name",
        "case_name",
    ),
    "priority": ("优先级", "priority", "级别"),
    "step": ("步骤", "步骤名称", "步骤描述", "step", "step name"),
}

STATUS_COLORS = {
    "pass": "C6EFCE",
    "通过": "C6EFCE",
    "成功": "C6EFCE",
    "✅": "C6EFCE",
    "fail": "FFC7CE",
    "失败": "FFC7CE",
    "不通过": "FFC7CE",
    "❌": "FFC7CE",
    "blocked": "FFEB9C",
    "阻塞": "FFEB9C",
    "⛔": "FFEB9C",
    "⚠": "FFEB9C",
    "待数据": "FFEB9C",
    "待测试": "FFEB9C",
    "待跑": "FFEB9C",
    "🟡": "FFEB9C",
    "🟢": "C6EFCE",
    "skip": "D9E1F2",
    "跳过": "D9E1F2",
    "⏭": "D9E1F2",
    "☑": "D9E1F2",
}


class AnnotationError(ValueError):
    """结果无法安全定位或输入契约不正确。"""


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def _norm(value: Any) -> str:
    """归一化字段值，只处理空白和大小写，不删除业务字符。"""

    return re.sub(r"\s+", "", _text(value)).casefold()


def _first(record: Mapping[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in record and record[key] not in (None, ""):
            return record[key]
    return None


def _as_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        number = int(value)
    except (TypeError, ValueError) as exc:
        raise AnnotationError(f"行号必须是整数，收到: {value!r}") from exc
    if number < 1:
        raise AnnotationError(f"行号必须大于 0，收到: {number}")
    return number


def _evidence_items(value: Any) -> list[tuple[str, str | None]]:
    if value in (None, ""):
        return []
    if isinstance(value, (str, Path)):
        return [(_text(value), _text(value))]
    if isinstance(value, Mapping):
        path = _first(value, "path", "file", "filepath", "uri")
        description = _first(value, "description", "text", "label")
        if path is not None and description:
            return [(f"{_text(description)}: {_text(path)}", _text(path))]
        text = _text(path or description)
        return [(text, _text(path) if path is not None else None)] if text else []
    if isinstance(value, Iterable):
        items: list[tuple[str, str | None]] = []
        for item in value:
            items.extend(_evidence_items(item))
        return items
    return [(_text(value), None)]


def _as_evidence(value: Any) -> list[str]:
    return [display for display, _ in _evidence_items(value) if display]


def _as_evidence_paths(value: Any) -> list[str]:
    return [path for _, path in _evidence_items(value) if path]


def _load_result_document(path: Path) -> Any:
    if not path.exists():
        raise AnnotationError(f"结果文件不存在: {path}")
    try:
        content = path.read_text(encoding="utf-8")
    except UnicodeDecodeError as exc:
        raise AnnotationError(f"结果文件必须使用 UTF-8 编码: {path}") from exc
    if path.suffix.casefold() in {".yaml", ".yml"}:
        return yaml.safe_load(content)
    try:
        return json.loads(content)
    except json.JSONDecodeError as exc:
        raise AnnotationError(f"结果文件不是合法 JSON: {path}: {exc}") from exc


def normalize_results(document: Any) -> list[dict[str, Any]]:
    """把常见的结果外壳归一化为结果记录列表。

    支持 ``[{...}]``、``{"cases": [{...}]}``、``{"results": [{...}]}``。
    记录内字段兼容 ``id/tc_id``、``name``、``result``、``time`` 等常见叫法。
    """

    if document is None:
        return []
    if isinstance(document, Mapping):
        payload = _first(document, "cases", "results", "items", "data")
        if payload is None:
            # 允许单条结果对象，减少 LLM 生成结果时的包裹要求。
            payload = [document]
    else:
        payload = document
    if not isinstance(payload, list):
        raise AnnotationError("结果必须是列表，或包含 cases/results/items/data 列表的对象")

    normalized: list[dict[str, Any]] = []
    for index, item in enumerate(payload, start=1):
        if not isinstance(item, Mapping):
            raise AnnotationError(f"第 {index} 条结果必须是对象，收到: {item!r}")
        record = dict(item)
        record["case_id"] = _text(_first(record, "case_id", "caseId", "id", "tc_id", "tcId")) or None
        record["case_name"] = _text(_first(record, "case_name", "caseName", "name", "用例名称")) or None
        record["sheet"] = _text(_first(record, "sheet", "sheet_name", "sheetName", "工作表")) or None
        record["row"] = _as_int(_first(record, "row", "row_number", "rowNumber", "source_row", "sourceRow"))
        record["status"] = _text(_first(record, "status", "state", "result_status", "判定")) or "·待跑"
        record["actual"] = _text(_first(record, "actual", "result", "actual_result", "实测结果", "description"))
        record["tier"] = _text(_first(record, "tier", "level", "档位"))
        record["tested_at"] = _text(_first(record, "tested_at", "testedAt", "time", "date", "测试时间"))
        raw_evidence = _first(record, "evidence", "evidence_paths", "evidencePaths", "screenshots", "证据")
        record["evidence"] = _as_evidence(raw_evidence)
        record["evidence_paths"] = _as_evidence_paths(raw_evidence)
        normalized.append(record)
    return normalized


def load_results(path: str | os.PathLike[str]) -> list[dict[str, Any]]:
    """从 JSON/YAML 结果文件读取并归一化结果。"""

    result_path = Path(path).expanduser().resolve()
    return normalize_results(_load_result_document(result_path))


def _header_kind(value: Any) -> str | None:
    normalized = _norm(value)
    if not normalized:
        return None
    for kind, aliases in HEADER_ALIASES.items():
        if normalized in {_norm(alias) for alias in aliases}:
            return kind
    return None


def _column_reference(value: str | int) -> int | None:
    """把列号、Excel 字母列或表头文字转换成列号。"""

    if isinstance(value, int):
        return value if value > 0 else None
    text = _text(value)
    if text.isdigit():
        number = int(text)
        return number if number > 0 else None
    if re.fullmatch(r"[A-Za-z]{1,3}", text):
        try:
            return column_index_from_string(text.upper())
        except ValueError:
            return None
    return None


def _column_from_reference(ws: Any, header_row: int, reference: str | int, kind: str) -> int:
    column_number = _column_reference(reference)
    if column_number is None:
        target = _norm(reference)
        for candidate in range(1, ws.max_column + 1):
            if _norm(ws.cell(header_row, candidate).value) == target:
                return candidate
        raise AnnotationError(f"工作表 {ws.title!r} 的第 {header_row} 行找不到 {kind} 列: {reference!r}")
    if column_number > ws.max_column and column_number > 16384:
        raise AnnotationError(f"工作表 {ws.title!r} 的 {kind} 列超出 Excel 范围: {reference!r}")
    return column_number


def _columns_from_header(ws: Any, header_row: int, *, case_id_column: str | int | None = None, case_name_column: str | int | None = None) -> dict[str, int]:
    columns: dict[str, int] = {}
    for column_number in range(1, ws.max_column + 1):
        kind = _header_kind(ws.cell(header_row, column_number).value)
        if kind and kind not in columns:
            columns[kind] = column_number
    if case_id_column is not None:
        columns["case_id"] = _column_from_reference(ws, header_row, case_id_column, "用例 ID")
    if case_name_column is not None:
        columns["case_name"] = _column_from_reference(ws, header_row, case_name_column, "用例名称")
    return columns


def find_header_row(ws: Any, scan_limit: int = 30, *, case_id_column: str | int | None = None, case_name_column: str | int | None = None, header_row: int | None = None) -> tuple[int | None, dict[str, int]]:
    """查找表头行，并返回 ``{字段类型: 列号}``。"""

    if header_row is not None:
        if header_row < 1 or header_row > ws.max_row:
            raise AnnotationError(f"工作表 {ws.title!r} 的表头行超出范围: {header_row}")
        return header_row, _columns_from_header(
            ws,
            header_row,
            case_id_column=case_id_column,
            case_name_column=case_name_column,
        )

    upper = min(ws.max_row, scan_limit)
    best: tuple[int | None, dict[str, int]] = (None, {})
    for row_number in range(1, upper + 1):
        columns = _columns_from_header(ws, row_number)
        if "case_name" in columns or "case_id" in columns:
            # 同时命中步骤/优先级的行更像真正表头；只有用例名也允许，适配极简用例表。
            score = len(columns) + (2 if "step" in columns else 0) + (1 if "priority" in columns else 0)
            previous_score = len(best[1]) + (2 if "step" in best[1] else 0) + (1 if "priority" in best[1] else 0)
            if best[0] is None or score > previous_score:
                best = (row_number, columns)
    return best


def _row_is_empty(ws: Any, row_number: int) -> bool:
    return all(ws.cell(row_number, column).value in (None, "") for column in range(1, ws.max_column + 1))


def _case_rows(ws: Any, header_row: int, columns: Mapping[str, int]) -> list[dict[str, Any]]:
    """读取数据行，并对合并/空白的用例名和 ID 做向下填充。"""

    rows: list[dict[str, Any]] = []
    current_id: str | None = None
    current_name: str | None = None
    for row_number in range(header_row + 1, ws.max_row + 1):
        if _row_is_empty(ws, row_number):
            continue
        raw_id = _text(ws.cell(row_number, columns["case_id"]).value) if "case_id" in columns else ""
        raw_name = _text(ws.cell(row_number, columns["case_name"]).value) if "case_name" in columns else ""
        if raw_id:
            current_id = raw_id
        if raw_name:
            if "case_id" in columns and not raw_id and current_name and _norm(raw_name) != _norm(current_name):
                # 新用例没有填写 ID 时，不把上一个用例的 ID 误带过来。
                current_id = None
            current_name = raw_name
        rows.append({"row": row_number, "case_id": current_id, "case_name": current_name})
    return rows


def _blocks(row_numbers: Sequence[int]) -> list[list[int]]:
    if not row_numbers:
        return []
    result: list[list[int]] = [[row_numbers[0]]]
    for row_number in row_numbers[1:]:
        if row_number == result[-1][-1] + 1:
            result[-1].append(row_number)
        else:
            result.append([row_number])
    return result


def _select_sheets(wb: Any, sheet_name: str | None) -> list[Any]:
    if not sheet_name:
        return list(wb.worksheets)
    if sheet_name not in wb.sheetnames:
        raise AnnotationError(f"结果指定的工作表不存在: {sheet_name}")
    return [wb[sheet_name]]


def _resolve_rows(wb: Any, record: Mapping[str, Any], case_rows: Mapping[str, list[dict[str, Any]]]) -> list[tuple[Any, int]]:
    sheet_name = _text(record.get("sheet")) or None
    row_number = record.get("row")
    case_id = _norm(record.get("case_id"))
    case_name = _norm(record.get("case_name"))
    if row_number is None and not case_id and not case_name:
        raise AnnotationError("结果至少需要 row/source_row、case_id 或 case_name 之一")

    candidates: list[tuple[Any, dict[str, Any]]] = []
    for ws in _select_sheets(wb, sheet_name):
        for row in case_rows.get(ws.title, []):
            if row_number is not None:
                if row["row"] == row_number:
                    candidates.append((ws, row))
            elif case_id:
                if _norm(row.get("case_id")) == case_id:
                    candidates.append((ws, row))
            elif _norm(row.get("case_name")) == case_name:
                candidates.append((ws, row))

    if row_number is not None:
        if not candidates:
            target = f"{sheet_name}!{row_number}" if sheet_name else str(row_number)
            raise AnnotationError(f"找不到结果指定的行: {target}")
        candidate_sheets = {ws.title for ws, _ in candidates}
        if not sheet_name and len(candidate_sheets) > 1:
            locations = ", ".join(f"{name}!{row_number}" for name in sorted(candidate_sheets))
            raise AnnotationError(f"行号在多个工作表中都存在，请补充 sheet；候选位置: {locations}")
        expanded: list[tuple[Any, int]] = []
        for ws, selected_row in candidates:
            # row 是最可靠的消歧键；同时给出 case_name 时，把该行扩展为同一
            # 连续用例区块，覆盖合并单元格/空白用例名下的后续步骤。
            identity_kind = None
            identity_value = ""
            if case_id and _norm(selected_row.get("case_id")) == case_id:
                identity_kind, identity_value = "case_id", case_id
            elif case_name and _norm(selected_row.get("case_name")) == case_name:
                identity_kind, identity_value = "case_name", case_name
            if identity_kind:
                all_rows = [
                    item["row"]
                    for item in case_rows.get(ws.title, [])
                    if _norm(item.get(identity_kind)) == identity_value
                ]
                block = next((block for block in _blocks(sorted(all_rows)) if selected_row["row"] in block), None)
                if block:
                    expanded.extend((ws, row) for row in block)
                    continue
            expanded.append((ws, selected_row["row"]))
        return expanded
    if not candidates:
        return []

    # 一个连续区块可代表一个多步骤用例；多个不连续区块说明名称/ID不唯一，
    # 必须让调用方补充 row，不能把结果写到可能是另一个用例的行上。
    grouped: dict[str, list[int]] = {}
    for ws, row in candidates:
        grouped.setdefault(ws.title, []).append(row["row"])
    blocks = [(name, block) for name, rows in grouped.items() for block in _blocks(sorted(rows))]
    if len(blocks) > 1:
        identity = record.get("case_id") or record.get("case_name")
        locations = ", ".join(f"{name}!{block[0]}-{block[-1]}" for name, block in blocks)
        raise AnnotationError(f"结果定位不唯一 ({identity!r})，请补充 row/source_row；候选区块: {locations}")
    return [(ws, row["row"]) for ws, row in candidates]


def _copy_header_style(source: Any, target: Any) -> None:
    if source is None:
        return
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.alignment = copy.copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy.copy(source.protection)


def _ensure_output_columns(ws: Any, header_row: int) -> dict[str, int]:
    existing: dict[str, int] = {}
    for column_number in range(1, ws.max_column + 1):
        value = _text(ws.cell(header_row, column_number).value)
        if value in OUTPUT_HEADERS and value not in existing:
            existing[value] = column_number
    next_column = ws.max_column + 1
    source_header = ws.cell(header_row, 1)
    for header in OUTPUT_HEADERS:
        if header not in existing:
            existing[header] = next_column
            cell = ws.cell(header_row, next_column, header)
            _copy_header_style(source_header, cell)
            cell.font = Font(name=cell.font.name, sz=cell.font.sz, bold=True, italic=cell.font.italic, color=cell.font.color)
            next_column += 1
    return existing


def _hide_legacy_columns(ws: Any, header_row: int) -> None:
    """兼容旧版输出：ID/档位不再显示在实际用例页，但保留数据供追溯。"""

    for column_number in range(1, ws.max_column + 1):
        if _text(ws.cell(header_row, column_number).value) in LEGACY_HIDDEN_HEADERS:
            ws.column_dimensions[get_column_letter(column_number)].hidden = True


def _status_fill(status: str) -> PatternFill | None:
    normalized = _norm(status)
    for prefix, color in STATUS_COLORS.items():
        if normalized.startswith(_norm(prefix)):
            return PatternFill(fill_type="solid", fgColor=color)
    return None


def _format_output_cell(cell: Any, *, wrap: bool = True) -> None:
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)


def _looks_like_path(value: str) -> bool:
    path = Path(value)
    return path.suffix.casefold() in IMAGE_SUFFIXES or "/" in value or "\\" in value or value.startswith(".")


def _resolve_evidence_path(raw: str, *, evidence_root: Path | None, result_dir: Path | None, source_dir: Path) -> Path | None:
    candidate = Path(raw).expanduser()
    candidates = [candidate] if candidate.is_absolute() else []
    if not candidate.is_absolute():
        for root in (evidence_root, result_dir, source_dir):
            if root:
                candidates.append(root / candidate)
    for item in candidates:
        if item.exists() and item.is_file():
            return item.resolve()
    return None


def _embed_evidence(ws: Any, cell: Any, evidence: Sequence[str], evidence_paths: Sequence[str] | None, *, evidence_root: Path | None, result_dir: Path | None, source_dir: Path, width: int, warnings: list[str]) -> None:
    if not evidence:
        return
    paths: list[Path] = []
    raw_paths = list(evidence_paths or evidence)
    for raw in raw_paths:
        path = _resolve_evidence_path(raw, evidence_root=evidence_root, result_dir=result_dir, source_dir=source_dir)
        if path and path.suffix.casefold() in IMAGE_SUFFIXES:
            paths.append(path)
        elif _looks_like_path(raw):
            warnings.append(f"证据文件不存在或不是图片，已保留文本: {raw}")
    if not paths:
        return
    # 同一结果重复执行时，不重复嵌入相同证据图片。
    if cell.value == "\n".join(evidence):
        return
    for path in paths:
        image = XLImage(str(path))
        if image.width and image.height:
            image.height = int(width * image.height / image.width)
        image.width = width
        ws.add_image(image, cell.coordinate)
        break


def _set_cell_value(cell: Any, value: Any, *, fill: PatternFill | None = None) -> None:
    cell.value = value if value not in (None, "") else None
    _format_output_cell(cell)
    if fill:
        cell.fill = fill


def _write_summary(wb: Any, report: Mapping[str, Any], source_path: Path, generated_at: str) -> None:
    if SUMMARY_SHEET in wb.sheetnames:
        del wb[SUMMARY_SHEET]
    ws = wb.create_sheet(SUMMARY_SHEET)
    ws.freeze_panes = "A2"
    ws.append(["字段", "值"])
    ws.append(["源文件", str(source_path)])
    ws.append(["生成时间", generated_at])
    ws.append(["匹配结果数", len(report["matched"])])
    ws.append(["未匹配结果数", len(report["unmatched"])])
    ws.append(["警告数", len(report["warnings"])])
    ws.append([])
    ws.append(["状态", "数量"])
    for status, count in sorted(Counter(item["status"] for item in report["matched"]).items()):
        ws.append([status, count])
    ws.append([])
    ws.append(["工作表", "行号", "用例 ID", "用例名称", "档位", "状态", "实测结果", "证据"])
    for item in report["matched"]:
        ws.append([
            item["sheet"],
            ",".join(str(row) for row in item["rows"]),
            item.get("case_id") or "",
            item.get("case_name") or "",
            item.get("tier") or "",
            item["status"],
            item["actual"],
            "\n".join(item["evidence"]),
        ])
    if report["unmatched"]:
        ws.append([])
        ws.append(["未匹配结果", "原因"])
        for item in report["unmatched"]:
            ws.append([item["identity"], item["reason"]])
    if report["warnings"]:
        ws.append([])
        ws.append(["警告"])
        for warning in report["warnings"]:
            ws.append([warning])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    for row in ws.iter_rows():
        for cell in row:
            _format_output_cell(cell)
    for column, width in {"A": 24, "B": 22, "C": 20, "D": 40, "E": 14, "F": 18, "G": 60, "H": 36}.items():
        ws.column_dimensions[column].width = width


def annotate_workbook(
    src: str | os.PathLike[str],
    results: Sequence[Mapping[str, Any]] | Mapping[str, Any],
    out: str | os.PathLike[str] | None = None,
    *,
    evidence_root: str | os.PathLike[str] | None = None,
    result_dir: str | os.PathLike[str] | None = None,
    sheet: str | None = None,
    header_row: int | None = None,
    case_id_column: str | int | None = None,
    case_name_column: str | int | None = None,
    generated_at: str | None = None,
    strict: bool = False,
    append_summary: bool = True,
    evidence_width: int = 150,
) -> dict[str, Any]:
    """回填结果并返回机器可读报告。

    ``results`` 可以直接是结果列表，也可以是 ``{"cases": [...]}`` 对象。
    常见中文/英文表头会自动识别；遇到自定义表头时传入 ``header_row`` 和
    ``case_id_column``/``case_name_column``，列参数可用列号、Excel 字母或表头文字。
    ``strict=True`` 时，只要存在未匹配结果就抛出 :class:`AnnotationError`；
    歧义定位无论 strict 与否都会抛错。
    """

    source_path = Path(src).expanduser().resolve()
    if source_path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise AnnotationError(f"仅支持 .xlsx/.xlsm 文件，不支持: {source_path.suffix or source_path.name}")
    if not source_path.exists():
        raise AnnotationError(f"用例文件不存在: {source_path}")
    if evidence_width < 20:
        raise AnnotationError("evidence_width 不能小于 20")

    output_path = Path(out).expanduser().resolve() if out else source_path.with_name(f"{source_path.stem}_AI自测结果{source_path.suffix}")
    if output_path == source_path:
        raise AnnotationError("输出文件不能覆盖源用例文件，请指定不同的 --out 路径")
    if isinstance(results, Mapping):
        result_records = normalize_results(results)
    else:
        result_records = normalize_results(list(results))
    result_root = Path(result_dir).expanduser().resolve() if result_dir else None
    evidence_root_path = Path(evidence_root).expanduser().resolve() if evidence_root else None
    generated = generated_at or datetime.now().astimezone().isoformat(timespec="seconds")

    keep_vba = source_path.suffix.casefold() == ".xlsm"
    wb = load_workbook(source_path, keep_vba=keep_vba, keep_links=True)
    case_rows: dict[str, list[dict[str, Any]]] = {}
    sheet_headers: dict[str, tuple[int, dict[str, int]]] = {}
    skipped_sheets: list[str] = []
    for ws in wb.worksheets:
        if ws.title == SUMMARY_SHEET:
            continue
        detected_header_row, columns = find_header_row(ws, header_row=header_row)
        if detected_header_row is None:
            skipped_sheets.append(ws.title)
            continue
        columns = _columns_from_header(
            ws,
            detected_header_row,
            case_id_column=case_id_column,
            case_name_column=case_name_column,
        )
        sheet_headers[ws.title] = (detected_header_row, columns)
        case_rows[ws.title] = _case_rows(ws, detected_header_row, columns)

    report: dict[str, Any] = {
        "source": str(source_path),
        "output": str(output_path),
        "matched": [],
        "unmatched": [],
        "warnings": [],
        "skipped_sheets": skipped_sheets,
    }

    for index, record in enumerate(result_records, start=1):
        resolve_record = record
        if sheet and not record.get("sheet"):
            resolve_record = dict(record)
            resolve_record["sheet"] = sheet
        matches = _resolve_rows(wb, resolve_record, case_rows)
        if not matches:
            identity = record.get("case_id") or record.get("case_name") or record.get("row") or f"结果#{index}"
            report["unmatched"].append({
                "identity": str(identity),
                "reason": "未找到对应工作表/用例 ID/用例名称/行号",
                "case_id": record.get("case_id"),
                "case_name": record.get("case_name"),
            })
            continue

        matched_by_sheet: dict[str, list[int]] = {}
        for ws, row_number in matches:
            matched_by_sheet.setdefault(ws.title, []).append(row_number)
        for ws_name, row_numbers in matched_by_sheet.items():
            ws = wb[ws_name]
            header_row, _ = sheet_headers[ws_name]
            _hide_legacy_columns(ws, header_row)
            output_columns = _ensure_output_columns(ws, header_row)
            _, source_columns = sheet_headers[ws_name]
            tier = record.get("tier")
            if not tier and "priority" in source_columns:
                tier = _text(ws.cell(min(row_numbers), source_columns["priority"]).value)
            for row_number in row_numbers:
                status_cell = ws.cell(row_number, output_columns["🤖AI状态"])
                actual_cell = ws.cell(row_number, output_columns["🤖AI实测结果"])
                evidence_cell = ws.cell(row_number, output_columns["🤖AI证据"])
                time_cell = ws.cell(row_number, output_columns["🤖AI时间"])
                _set_cell_value(status_cell, record.get("status"), fill=_status_fill(_text(record.get("status"))))
                _set_cell_value(actual_cell, record.get("actual"))
                evidence_text = "\n".join(record.get("evidence", []))
                if row_number == min(row_numbers):
                    _embed_evidence(
                        ws,
                        evidence_cell,
                        record.get("evidence", []),
                        record.get("evidence_paths", []),
                        evidence_root=evidence_root_path,
                        result_dir=result_root,
                        source_dir=source_path.parent,
                        width=evidence_width,
                        warnings=report["warnings"],
                    )
                _set_cell_value(evidence_cell, evidence_text)
                _set_cell_value(time_cell, record.get("tested_at") or generated)
                ws.column_dimensions[get_column_letter(output_columns["🤖AI状态"])].width = 18
                ws.column_dimensions[get_column_letter(output_columns["🤖AI实测结果"])].width = 58
                ws.column_dimensions[get_column_letter(output_columns["🤖AI证据"])].width = 34
                ws.column_dimensions[get_column_letter(output_columns["🤖AI时间"])].width = 22
                if row_number == min(row_numbers) and evidence_text:
                    height = ws.row_dimensions[row_number].height or 15
                    ws.row_dimensions[row_number].height = max(height, min(260, evidence_width * 2.25))
            report["matched"].append({
                "sheet": ws_name,
                "rows": sorted(row_numbers),
                "case_id": record.get("case_id"),
                "case_name": record.get("case_name"),
                "tier": tier,
                "status": record.get("status"),
                "actual": record.get("actual"),
                "evidence": record.get("evidence", []),
            })

    if strict and report["unmatched"]:
        details = ", ".join(item["identity"] for item in report["unmatched"])
        raise AnnotationError(f"存在未匹配结果，strict 模式拒绝保存: {details}")
    if append_summary:
        _write_summary(wb, report, source_path, generated)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    report["generated_at"] = generated
    return report


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="将 JSON/YAML 自测结果回填到任意 Excel 用例文件")
    parser.add_argument("--src", required=True, help="源用例文件，支持 .xlsx/.xlsm")
    result_group = parser.add_mutually_exclusive_group(required=True)
    result_group.add_argument("--results", help="结果 JSON/YAML 文件")
    result_group.add_argument("--results-json", help="直接传入结果 JSON 字符串")
    parser.add_argument("--out", help="输出文件；默认写到源文件旁的 *_AI自测结果.xlsx")
    parser.add_argument("--evidence-root", help="相对证据路径的根目录")
    parser.add_argument("--sheet", help="仅处理指定工作表；结果记录中的 sheet 优先")
    parser.add_argument("--header-row", type=int, help="自定义表头行号；未提供时自动识别常见表头")
    parser.add_argument("--case-id-column", help="自定义用例 ID 列：列号、Excel 字母或表头文字")
    parser.add_argument("--case-name-column", help="自定义用例名称列：列号、Excel 字母或表头文字")
    parser.add_argument("--date", help="统一回填时间；不传则使用当前本地时间")
    parser.add_argument("--evidence-width", type=int, default=150, help="嵌入证据图片宽度，默认 150px")
    parser.add_argument("--strict", action="store_true", help="有未匹配结果时拒绝保存")
    parser.add_argument("--no-summary", action="store_true", help="不生成 🤖AI自测汇总 工作表")
    return parser


def _configure_stdio() -> None:
    """Windows 默认代码页可能是 GBK，CLI 输出的 AI 列名含 emoji。"""

    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if reconfigure:
            reconfigure(encoding="utf-8", errors="replace")


def main(argv: Sequence[str] | None = None) -> int:
    _configure_stdio()
    args = _build_parser().parse_args(argv)
    try:
        if args.results:
            result_path = Path(args.results).expanduser().resolve()
            document = _load_result_document(result_path)
            result_dir = result_path.parent
        else:
            try:
                document = json.loads(args.results_json)
            except json.JSONDecodeError as exc:
                raise AnnotationError(f"--results-json 不是合法 JSON: {exc}") from exc
            result_dir = None
        report = annotate_workbook(
            args.src,
            document,
            args.out,
            evidence_root=args.evidence_root,
            result_dir=result_dir,
            sheet=args.sheet,
            header_row=args.header_row,
            case_id_column=args.case_id_column,
            case_name_column=args.case_name_column,
            generated_at=args.date,
            strict=args.strict,
            append_summary=not args.no_summary,
            evidence_width=args.evidence_width,
        )
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return 0
    except (AnnotationError, OSError, TypeError, yaml.YAMLError) as exc:
        print(f"annotate_excel: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
